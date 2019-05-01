import codecs
import csv
import os
import sys

import openpyxl
import pymysql
import xlrd
import xlwt
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from xlrd import open_workbook

from fileopen import *
from show_graph import *
from Ui_1130_1 import *
from Ui_Add_2 import *
from Ui_sample_info import *
from Ui_sec_win import *


class sec_win(QtWidgets.QMainWindow, Ui_Dialog):
    def __init__(self,parent=None):
        super(sec_win, self).__init__(parent)
        self.setupUi(self)
        self.treeWidget.expandAll()
        self.setWindowIcon(QIcon(('./piliang_ico/Toolsnet.png')))
        self.setWindowTitle('学生管理 Client')
        self.setFixedSize(self.width(),self.height())
        self.reflash_table()
        self.treeWidget.doubleClicked.connect(self.item_reaction)
        

    def reflash_table(self):
        self.info=self.get_score_info_from_db()
        self.stu_id=[x[0] for x in self.info]
        self.tableWidget.setRowCount(len(self.info))
        for i in range(len(self.info)):
            for j in range(len(self.info[i])):
                self.tableWidget.setItem(i,j,QTableWidgetItem(str(self.info[i][j])))

    def get_item(self):
        Item_list=self.treeWidget.selectedItems()  #return the selected item as a list
        for ii in Item_list:
            print(ii.text(0))
            return ii.text(0)
            continue

    def item_reaction(self):
        item=self.get_item()
        if item=='刷新':
            self.reflash_table()
        elif item=='按学号':
            self.D2=Dialog2()
            self.D2.show()
            self.D2.pushButton.clicked.connect(self.select_byid)
            self.D2.pushButton.clicked.connect(self.D2.close)
        elif item=='按姓名':
            self.D1=Dialog1()
            self.D1.show()
            self.D1.pushButton.clicked.connect(self.select_byname)
            self.D1.pushButton.clicked.connect(self.D1.close)
        elif item=='大学物理':
            self.show_phy=show_pic('大学物理')
            self.show_phy.show()
        elif item=='离散数学':
            self.show_dis=show_pic('离散数学')
            self.show_dis.show()
        elif item=='线性代数':
            self.show_liner=show_pic('线性代数')
            self.show_liner.show()
        elif item=='平均成绩':
            self.show_ave=show_pic('平均成绩')
            self.show_ave.show()
        elif item=='手动添加':
            self.add_1=Add_d2()
            self.add_1.show()
            self.add_1.pushButton.clicked.connect(self.Add_1_reaction)
        elif item=='文件导入':
            # add_2=fileopen_2()
            # self.filevalues=add_2.read_file()
            # self.Add_2_reaction()
            # self.reflash_table()
            def func():
                add_2=fileopen_2()
                self.filevalues=add_2.read_file()
                self.Add_2_reaction()
                self.reflash_table()
            # QMessageBox.information(self,"提示!","目前仅支持从*.txt / *.csv文件导入,并且确保文件格式正确!")
            self.sample_info=sample('./piliang_ico/sample_txt2.JPG','./piliang_ico/sample_csv2.JPG')
            self.sample_info.show()
            self.sample_info.pushButton.clicked.connect(func)
        elif item=='导出为 */.txt':
            self.output_txt()
        elif item=='导出为 */.csv':
            self.output_csv()
        elif item=='导出为 */.xlsx':
            self.output_xlsx()

    def Add_1_reaction(self):
        self.info=self.get_score_info_from_db()   #刷新但不调用reflash_table(),优化时间
        self.stu_id=[x[0] for x in self.info]     #
        stu_id=''.join(self.add_1.lineEdit_1.text().split())
        name=self.add_1.lineEdit_2.text()
        phy=self.add_1.lineEdit_3.text()
        dis=self.add_1.lineEdit_4.text()
        liner=self.add_1.lineEdit_5.text()
        if stu_id not in self.stu_id:
            if stu_id!='' and name!='' and phy!='' and dis!='' and liner!='':
                if phy.isdigit() and dis.isdigit() and liner.isdigit():
                    if int(phy)<100 and int(phy)>0 and int(dis)<100 and int(dis)>0 and int(liner)<100 and int(liner)>0:
                        ave=round((int(phy)+int(dis)+int(liner))/3,1)
                        db=pymysql.connect(
                            host='localhost',
                            user='root',
                            password='Achencan123',
                            db='test',
                            port=3306
                        )
                        cur=db.cursor()
                        sql_pattern="""
                        insert into score_info values(
                            '{}',
                            '{}',
                            '{}',
                            '{}',
                            '{}',
                            '{}'
                        )
                        """
                        try:
                            sql_insert=sql_pattern.format(stu_id,name,phy,dis,liner,ave)
                            cur.execute(sql_insert)
                            db.commit()
                        except Exception as e:
                            print(e)
                            db.rollback()
                        finally:
                            db.close()
                            self.add_1.close()
                            self.reflash_table()
                            QMessageBox.information(self,"提示!","Success!")

                        pass
                    else:
                        QMessageBox.warning(self,"警告!","请检查成绩范围(0~100)")
                        self.add_1.raise_()
                else:
                    QMessageBox.warning(self,"警告!","各项成绩应该为整数!")
                    self.add_1.raise_()
            else:
                QMessageBox.warning(self,"警告!","请填写完整信息!")
                self.add_1.raise_()
        else:
            QMessageBox.warning(self,"警告!","此人成绩已存在,请勿重复录入!")
            self.add_1.raise_()

    def Add_2_reaction(self):
        if self.filevalues is not None and self.filevalues!=[]:
            db=pymysql.connect(
                host='localhost',
                user='root',
                password='Achencan123',
                db='test',
                port=3306
            )
            cur=db.cursor()
            sql_pattren="""
                insert into score_info values(
                    '{}',
                    '{}',
                    '{}',
                    '{}',
                    '{}',
                    '{}'
                )
            """

            if len(self.filevalues[0])%6==0 and self.filevalues[0][0].isdigit() and not self.filevalues[0][1].isdigit():
                try:
                    count1=0
                    for each in self.filevalues:
                        stu_id=str(each[0])
                        if stu_id not in self.stu_id:
                            name=str(each[1])
                            phy=str(each[2])
                            dis=str(each[3])
                            liner=str(each[4])
                            ave=str(round((int(phy)+int(dis)+int(liner))/3,1))
                            cur.execute(sql_pattren.format(stu_id,name,phy,dis,liner,ave))
                            db.commit()
                            count1+=1
                    count2=len(self.filevalues)-count1
                    QMessageBox.information(self,"提示!","成功导入{}条数据！其中{}条数据被过滤(原因:数据重复)".format(str(count1),str(count2)))
                except Exception as e:
                    print(e)
                    db.rollback()
                    QMessageBox.warning(self,"警告！","文件格式错误！请检查文件格式和编码格式是否为‘UTF-8’")
                finally:
                    db.close()
            else:
                QMessageBox.warning(self,"警告！","文件格式错误！请检查文件格式和编码格式是否为‘UTF-8’")
        

    def select_byid(self):
        self.linetext=self.D2.get_text().strip()
        for i in self.info:
            if i[0]==self.linetext:
                self.tableWidget.setRowCount(1)
                for j in range(6):
                    self.tableWidget.setItem(0,j,QTableWidgetItem(i[j]))
                break
        else:
            QMessageBox.warning(self,"提示",'未查找到此学号!')
        
    def select_byname(self):
        self.linetext=self.D1.get_text().strip()
        list_1=[]
        for i in self.info:
            if i[1]==self.linetext:
                list_1.append(i)
        if len(list_1)!=0:
            self.tableWidget.setRowCount(len(list_1))
            flag=0
            for i in list_1:
                for j in range(6):
                    self.tableWidget.setItem(flag,j,QTableWidgetItem(i[j]))
                flag+=1       
        else:
            QMessageBox.warning(self,"提示","未查找到此姓名!")

    def get_score_info_from_db(self):
        db=pymysql.connect(
            host='localhost',
            user='root',
            password='Achencan123',
            db='test',
            port=3306
        )
        cur=db.cursor()
        try:
            cur.execute("select * from score_info")
            results=cur.fetchall()
            return results
        except Exception as e:
            raise e
            db.rollback()
        finally:
            db.close()

    def output_txt(self):
        db=pymysql.connect(
            host='localhost',
            user='root',
            password='Achencan123',
            db='test',
            port=3306
        )
        cur=db.cursor()
        try:
            cur.execute("select * from score_info")
            data=cur.fetchall()
            with open(r'./导出文件/学生成绩.txt','w',encoding='utf-8',errors='ignore') as output:
                for each in data:
                    for i in each:
                        output.write(i+'\n')
            QMessageBox.information(self,"提示!","*/.txt文件导出成功！文件位置为\n{}".format(os.getcwd()+'\\导出文件\\学生成绩.txt'))
        except Exception as e:
            raise e
            QMessageBox.warning(self,"警告!","导出失败!")
        finally:
            db.close()
         
    def output_csv(self):
        db=pymysql.connect(
            host='localhost',
            user='root',
            password='Achencan123',
            db='test',
            port=3306
        )
        cur=db.cursor()
        try:
            cur.execute("select * from score_info")
            score=cur.fetchall()
            with open('./导出文件/学生成绩.csv','w',newline='') as output:
                writer=csv.writer(output)
                writer.writerow(['学号','姓名','大学物理','离散数学','线性代数','平均成绩'])
                for each in score:
                    writer.writerow(each)
            QMessageBox.information(self,"提示!","*/.csv文件导出成功！文件位置为\n{}".format(os.getcwd()+'\\导出文件\\学生成绩.csv'))
        except Exception as e:
            raise e
            QMessageBox.warning(self,"警告!","导出失败!")
        finally:
            db.close()

    def output_xlsx(self):
        db=pymysql.connect(
        host='localhost',
        user='root',
        password='Achencan123',
        db='test',
        port=3306
        )
        cur=db.cursor()
        try:
            cur.execute("select * from score_info")
            score=cur.fetchall()
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            title=[('学号','姓名','大学物理','离散数学','线性代数','平均成绩')]
            score=title+list(score)
            for row in range(1,len(score)+1):
                for col in range(1,7):
                    outws.cell(row, col).value = score[row-1][col-1]  # 写文件
            saveExcel = "./导出文件/学生成绩.xlsx"
            outwb.save(saveExcel)  # 一定要记得保存
            QMessageBox.information(self,"提示!","*/.xlsx文件导出成功！文件位置为\n{}".format(os.getcwd()+'\\导出文件\\学生成绩.xlsx'))
        except Exception as e:
            QMessageBox.warning(self,"警告!","导出失败!")
            print(e)
        finally:
            db.close()


# app = QtWidgets.QApplication(sys.argv)
# win=sec_win()
# win.show()
# sys.exit(app.exec_())



# with open('info.txt','w',encoding='gbk',errors='ignore') as info:
#     for i in range(1000):
#         list_1=['cchan'+str(i),str(631707060605+i),'23','男']
#         for x in list_1:
#             info.write(x+'\n')


 
# class tablewidget(QtWidgets.QTableWidget):
#     def __init__(self,parent=None):
#         super(tablewidget, self).__init__(parent)
#     def contextMenuEvent(self, event):
#         hitIndex = self.indexAt(event.pos()).row()
#         print(hitIndex)
#         if hitIndex > -1:
#             pmenu = QMenu(self)
#             pDeleteAct = QAction("删除",pmenu)
#             pmenu.addAction(pDeleteAct)
#             # pDeleteAct.triggered.connect(self.deleteItemSlot)
#             pmenu.popup(self.mapToGlobal(event.pos()))


# import pymysql
# from random import choice
# with open('score.txt','a+',encoding='utf-8') as score:
#     for i in range(1,1001):
#         score.write(str(621707060602+i)+'\n')
#         score.write('cchan'+str(i)+'\n')
#         score.write(str(choice(list(range(0,100))))+'\n')
#         score.write(str(choice(list(range(0,100))))+'\n')
#         score.write(str(choice(list(range(0,100))))+'\n')
#         score.write(str(choice(list(range(0,100))))+'\n')

# with open('score.txt','r',encoding='utf-8') as score:
#     info=[x.strip() for x in score.readlines()]
#     info=[info[i:i+6] for i in range(0,len(info)-6,6)]
#     db=pymysql.connect(
#         host='localhost',
#         user='root',
#         password='Achencan123',
#         db='test',
#         port=3306
#     )
#     cur=db.cursor()
#     sql_pattern="""
#     insert into score_info values(
#         '{}',
#         '{}',
#         '{}',
#         '{}',
#         '{}',
#         '{}'
#     );
#     """
#     try:
#         for each in info:
#             stu_id=each[0]
#             name=each[1]
#             phy=each[2]
#             dis=each[3]
#             liner=each[4]
#             ave=each[5]
#             sql_insert = sql_pattern.format(stu_id,name,phy,dis,liner,ave)
#             cur.execute(sql_insert)
#             db.commit()
#     except Exception as e:
#         raise e
#         db.rollback()
#     finally:
#         db.close()

# import csv
# with open('score.txt','r',) as score:
#     score=[x.strip() for x in score.readlines()]
#     score=[x for x in score]
#     score=[score[i:i+6] for i in range(0,len(score)-6,6)]
#     with open('./导出文件/学生成绩.csv','w',newline='') as output:
#         writer=csv.writer(output)
#         writer.writerow(['学号','姓名','大学物理','离散数学','线性代数','平均成绩'])
#         for each in score:
#             writer.writerow(each)
