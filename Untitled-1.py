import csv
from random import choice, randint, sample

import openpyxl
import pymysql
import xlrd
import xlwt
from pyecharts import Bar, Geo, Map, Pie
from xlrd import open_workbook



print(''.join(x.split()))

# with open('名单.txt','r') as score:
#     score=[x.strip() for x in score.readlines()]
#     score=[x.split('\t') for x in score]
#     # score=[x for x in score]
#     # score=[score[i:i+6] for i in range(0,len(score)-6,6)]
#     print(score)
#     with open('score.txt','w',encoding='utf-8') as out:
#         for i in score:
#             out.write(i[0]+'\n')
#             out.write(i[1]+'\n')
#             phy=randint(20,100)
#             dis=randint(20,100)
#             liner=randint(20,100)
#             out.write(str(phy)+'\n')
#             out.write(str(dis)+'\n')
#             out.write(str(liner)+'\n')
#             out.write(str(round((phy+dis+liner)/3, 1))+'\n')


# db=pymysql.connect(
#     host='localhost',
#     user='root',
#     password='Achencan123',
#     db='test',
#     port=3306
# )
# cur=db.cursor()
# try:
#     cur.execute("select * from score_info")
#     score=cur.fetchall()
# except Exception as e:
#     raise e
# finally:
#     db.close()
# outwb = openpyxl.Workbook()  # 打开一个将写的文件
# outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
# title=[('学号','姓名','大学物理','离散数学','线性代数','平均成绩')]
# score=title+list(score)
# print(score)
# for row in range(1,len(score)+1):
#     for col in range(1,7):
#         outws.cell(row, col).value = score[row-1][col-1]  # 写文件
# saveExcel = "./导出文件/学生成绩.xlsx"
# outwb.save(saveExcel)  # 一定要记得保存

# import codecs
# import csv
# with codecs.open('学生成绩.csv','r') as info:
#     info=csv.reader(info)
#     info=[x for x in info]
#     print(info)
